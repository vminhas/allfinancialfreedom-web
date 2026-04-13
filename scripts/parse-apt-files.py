#!/usr/bin/env python3
"""
Parse individual agent APT Excel files to extract phase checklist completion.
Reads all agent APT files from the downloaded Drive folder, extracts the
System sheet checklist, and outputs JSON keyed by agent code.

Usage: python3 scripts/parse-apt-files.py
Output: scripts/apt-checklist-data.json
"""

import openpyxl
import json
import os
import re
from datetime import datetime

APT_FOLDER = '/Users/melineeminhas/Downloads/drive-download-20260413T204049Z-3-001'
OUTPUT = '/Users/melineeminhas/dev/allfinancialfreedom-web/scripts/apt-checklist-data.json'

# Files to skip (templates, non-agent, duplicates)
EXCLUDE = {
    'APT Master Copy (Mandy).xlsx', 'APT Master Copy.xlsx', 'Sample System Progressions APT.xlsx',
    'Blank Business Partners_Agent Progress Tracker.xlsx', 'System Progressions GFI.xlsx',
    '(Share)System Progressions GFI.xlsx', ' System Progressions APT_.xlsx',
    '2024 4Buckets - Master Copy.xlsx', '4 BUCKETS_.xlsx', '4 Buckets + Dime.xlsx',
    'Copy of 4 Buckets   Dime.xlsx', 'ATTENDANCE 2023.xlsx', 'Attendance_.xlsx',
    'BUSINESS TRACKING TEMPLATE.xlsx', 'TEAM LEGACY_ BUSINESS TRACKING TEMPLATE.xlsx',
    'CFT Academy - Videos Trainings.xlsx', 'CFT in Progress.xlsx',
    'Lead Accountability Form (Responses).xlsx', 'EMD Policy Pipeline Master Copy_.xlsx',
    'Empower Policy Tracker.xlsx', 'Empower Training Schedule.xlsx',
    'March 2025_Cousins Napa_Sonoma Valley.xlsx', "Vick_s Logins & Info.xlsx",
    'GFI - ZITTT Agents.xlsx', 'GFI 4 - P I L L A R S (BKG).xlsx',
    'Hierarchy Tracker-Submission_.xlsx', 'Copy of MASTER - GFI Weekly Accountability Tracker.xlsx',
    'Copy of Rebranding list.xlsx', 'Divine Reach - Vick Minhas(1).xlsx',
    'Divine Reach - Vick Minhas.xlsx', 'ZT3 Internacional - PES Master Copy.xlsx',
    'ZT3 Leads Division Schedule 2026.xlsx', 'ZT3 National Team Training Schedule.xlsx',
    'ZT3 Recognition.xlsx', 'Team Empower Recognitions (9-28 Feb).xlsx',
    'Nov-De Empower Team Report 2025.xlsx', 'PropHog_08-09-2025_11-51-03.xlsx',
    'Extra Hours (Natalia Serrano).xlsx', 'MATT -  Rebranding and Field Training List.xlsx',
    'Zachs Top 25 List for Rebranding and Field Training.xlsx',
    'Jacob Denning Rebranding and Field Training List.xlsx',
    'Copy of Aaron Taylor System Progressions APT_.xlsx',
    'Christopher Simpson Agent Name System Progressions APT.xlsx',
    'George Haas System Progressions APT(1).xlsx',
    'Micah Standish  System Progressions APT_.xlsx',
    'Maren Sanders System Progressions APT.xlsx',
    'ZT3 2026 National Base, Licensing, and CFT.xlsx',
    'Clients.xlsx', 'Bryon Dahle - Policy Breakdown.xlsx',
    'Mercedes Grubb_s FTA Quick Notes.xlsx', 'Agent License Tracker.xlsx',
    'Wednesday Team Attendance.xlsx', 'KATE SUSSMAN APT.xlsx',
}

def is_checked(val):
    """Treat True, 1, 'x', 'yes', '✓', '✔' as checked"""
    if val is None: return False
    if isinstance(val, bool): return val
    if isinstance(val, (int, float)): return val > 0
    s = str(val).strip().lower()
    return s in ('true', '1', 'x', 'yes', 'y', '✓', '✔', 'done', 'complete', 'completed')

def get_cell(rows, row_idx, col_idx):
    """Safely get cell value"""
    if row_idx >= len(rows): return None
    row = rows[row_idx]
    if col_idx >= len(row): return None
    return row[col_idx]

def parse_system_sheet(ws):
    """
    Parse the System sheet to extract:
    - agent_code, name from row 2
    - checklist items from the grid starting row 14

    Grid layout (0-indexed rows, 0-indexed cols):
    Phase 1: labels col 0, checkboxes col 2  (rows 13-26)
    Phase 2: labels col 5, checkboxes col 7  (rows 13-32)
    Phase 3: labels col 8, checkboxes col 10 (rows 13-27)
    Phase 4: labels col 11, checkboxes col 13 (rows 13-32)
    Phase 5: labels col 14, checkboxes col 16/17 (rows 13-33)
    """
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 6:
        return None

    # Extract agent code from row 2 (0-indexed row 1), col 7
    agent_code = None
    agent_name = None
    if len(rows) > 1:
        row2 = rows[1]
        if len(row2) > 7 and row2[7]:
            agent_code = str(row2[7]).strip()
        if len(row2) > 0 and row2[0]:
            agent_name = str(row2[0]).strip()

    # Map our itemKeys to (row_index, col_index) in the spreadsheet
    # Rows are 0-indexed (row 14 in Excel = index 13)
    ITEM_MAP = {
        # Phase 1 — checkbox col 2
        'week1_onboarding':       (13, 2),
        'licensing_class':        (14, 2),
        'pfr':                    (15, 2),
        'fast_start_school':      (16, 2),
        'week2_onboarding':       (18, 2),
        '3_business_partners':    (19, 2),
        'business_marketing_plan':(20, 2),
        'pass_license_test':      (21, 2),
        'fingerprints_apply':     (22, 2),
        'week3_onboarding':       (24, 2),
        'master_scripts':         (25, 2),
        'schedule_10_trainings':  (26, 2),
        # Phase 2 — checkbox col 7
        'fta_1':                  (14, 7),
        'fta_2':                  (15, 7),
        'fta_3':                  (16, 7),
        'fta_4':                  (17, 7),
        'fta_5':                  (18, 7),
        'fta_6':                  (19, 7),
        'fta_7':                  (20, 7),
        'fta_8':                  (21, 7),
        'fta_9':                  (22, 7),
        'fta_10':                 (23, 7),
        'associate_promotion':    (24, 7),
        'direct_1':               (25, 7),
        'direct_2':               (26, 7),
        'direct_3':               (27, 7),
        'client_1':               (28, 7),
        'client_2':               (29, 7),
        'client_3':               (30, 7),
        'net_license':            (31, 7),
        'first_1000':             (32, 7),
        # Phase 3 — checkbox col 10
        'cft_classes':            (13, 10),
        'trainer_signoff':        (14, 10),
        'cft_coordinator_signoff':(15, 10),
        'emd_signoff':            (16, 10),
        'client_1st_apt':         (18, 10),
        'client_2nd_apt':         (19, 10),
        'phone_call_scripts':     (20, 10),
        'recruiting_interview':   (21, 10),
        'top_5_products':         (22, 10),
        # Phase 4 — checkbox col 13
        '45k_points':             (13, 13),
        'month1_premium':         (14, 13),
        'month2_premium':         (15, 13),
        'month3_premium':         (16, 13),
        'license_1':              (18, 13),
        'license_2':              (19, 13),
        'license_3':              (20, 13),
        'license_4':              (21, 13),
        'license_5':              (22, 13),
        # Phase 5 — checkbox col 16 (net points + MD) or col 17 (licenses)
        '150k_net_6mo':           (13, 16),
        '1_marketing_director':   (16, 16),
    }

    # Phase 5 licenses use col 17 (rows 18-37 for licenses 1-20)
    for i in range(1, 21):
        ITEM_MAP[f'license_{i}'] = (17 + i, 17)  # rows 18-37 → indexes 17-36... let me adjust
    # Actually from the data: License 1 is at row 19 (index 18), col 15
    # Let me use the col 15 for p5 licenses based on what I saw in Kiirah's file
    # Row 19 index 18: (14, 'License 1'), col 15 = True
    for i in range(1, 11):
        ITEM_MAP[f'p5_license_{i}'] = (17 + i, 15)  # rows 19-28

    checklist = {}
    for item_key, (row_i, col_i) in ITEM_MAP.items():
        val = get_cell(rows, row_i, col_i)
        checklist[item_key] = is_checked(val)

    # Fix: Phase 5 licenses use our DB key 'license_1' through 'license_20'
    # but Phase 4 also uses 'license_1' through 'license_5'
    # We already mapped phase 4 above; phase 5 licenses need separate handling
    # since they share key names. Phase 5 licenses:
    for i in range(1, 21):
        row_i = 18 + i - 1  # license 1 = row 19 = index 18
        # col 15 for first 10, col 17 for licenses 11-20
        col_i = 15 if i <= 10 else 17
        actual_row = 18 + i - 1
        if i > 10:
            actual_row = 18 + (i - 11)
        val = get_cell(rows, actual_row, col_i)
        checklist[f'p5_license_{i}'] = is_checked(val)

    return {
        'agentCode': agent_code,
        'agentName': agent_name,
        'checklist': checklist,
    }

# Process all agent files
results = {}
skipped = []
no_system_sheet = []

files = sorted(f for f in os.listdir(APT_FOLDER)
               if f.endswith('.xlsx') and f not in EXCLUDE and not f.startswith('~$'))

print(f'Processing {len(files)} files...')

for filename in files:
    filepath = os.path.join(APT_FOLDER, filename)
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        if 'System' not in wb.sheetnames:
            no_system_sheet.append(filename)
            wb.close()
            continue

        data = parse_system_sheet(wb['System'])
        wb.close()

        if not data or not data['agentCode']:
            skipped.append(f'{filename} (no agent code)')
            continue

        code = data['agentCode']
        # Deduplicate — keep the file with more checked items
        if code in results:
            existing_checked = sum(1 for v in results[code]['checklist'].values() if v)
            new_checked = sum(1 for v in data['checklist'].values() if v)
            if new_checked > existing_checked:
                results[code] = data
        else:
            results[code] = data

    except Exception as e:
        skipped.append(f'{filename} (error: {e})')

print(f'\n✓ Parsed {len(results)} unique agents')
print(f'  No System sheet: {len(no_system_sheet)}')
print(f'  Skipped: {len(skipped)}')
if no_system_sheet:
    print('  Files without System sheet:', no_system_sheet[:10])
if skipped:
    print('  Skipped files:', skipped[:10])

# Stats on checklist completion
total_items = sum(sum(1 for v in d['checklist'].values() if v)
                  for d in results.values())
print(f'\n  Total checked items across all agents: {total_items}')

with open(OUTPUT, 'w') as f:
    json.dump(results, f, indent=2)

print(f'\n✓ Written to {OUTPUT}')
