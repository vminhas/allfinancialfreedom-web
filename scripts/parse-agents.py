#!/usr/bin/env python3
"""
Parse AFF agent data from Excel spreadsheet into JSON for the seed script.
Usage: python3 scripts/parse-agents.py
Output: scripts/agents-seed-data.json
"""

import openpyxl
import json
import re
from datetime import datetime

XLSX_PATH = '/Users/melineeminhas/Downloads/New 2025 GFI - ZTT Agents.xlsx'

# Map spreadsheet carrier columns (Carrier Appointments sheet) → our carrier names
CARRIER_SHEET_MAP = {
    5: 'ANICO Life',
    6: 'ANICO Annuity',
    7: 'Augustar',
    8: 'Corebridge Annuity',
    9: 'Corebridge Life',
    10: 'Lincoln',
    11: 'Foresters',
    12: 'Mutual of Omaha',
    13: 'SILAC',
    14: 'American Equity',
    # col 15 = "North American" — maps to both Life and Annuity
    16: 'Equitrust',
    17: 'Prudential',
}

# Map Empower Tracker producer # columns → our carrier names
PRODUCER_COL_MAP = {
    33: ('American Equity', None),       # value IS the producer number
    34: ('Augustar', None),
    36: ('Corebridge Life', None),
    38: ('Corebridge Annuity', None),
    39: ('F&G Life', None),
    40: ('F&G Annuity', None),
    43: ('Mutual of Omaha', None),
    44: ('SILAC', None),
    45: ('North American Life', None),
    46: ('North American Annuity', None),
}

STATUS_MAP = {
    'appointed': 'APPOINTED',
    'pending request': 'PENDING',
    'j-i-t carrier': 'JIT',
    'jit': 'JIT',
}

def fmt_date(val):
    if not val:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    return str(val).strip() or None

def clean_str(val):
    if not val:
        return None
    s = str(val).strip()
    return s if s else None

def clean_npn(val):
    if not val:
        return None
    # NPN is a number, strip decimals
    s = str(val).strip().split('.')[0]
    return s if s else None

def split_name(full_name):
    if not full_name:
        return '', ''
    parts = str(full_name).strip().split(' ', 1)
    return parts[0], parts[1] if len(parts) > 1 else ''

wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)

# ── 1. Parse "Copy of Empower Tracker" for phase + goal + trainer ──────────
phase_data = {}   # agentCode → {phase, goal, trainer}
ws_copy = wb['Copy of Empower Tracker']
for row in list(ws_copy.iter_rows(values_only=True))[1:]:
    code = clean_str(row[3])
    if not code:
        continue
    phase = int(row[12]) if row[12] and str(row[12]).replace('.0','').isdigit() else 1
    phase_data[code] = {
        'phase': min(max(phase, 1), 5),
        'goal': clean_str(row[14]),
        'trainer': clean_str(row[15]),
    }

# ── 2. Parse "Carrier Appointments" sheet ─────────────────────────────────
carrier_data = {}  # agentCode → {carrierName: {status, producerNumber}}
ws_carriers = wb['Carrier Appointments']
for row in list(ws_carriers.iter_rows(values_only=True))[1:]:
    code = clean_str(row[1])
    if not code:
        continue
    if code not in carrier_data:
        carrier_data[code] = {}

    for col_idx, carrier_name in CARRIER_SHEET_MAP.items():
        val = clean_str(row[col_idx]) if len(row) > col_idx else None
        if val:
            status = STATUS_MAP.get(val.lower(), 'APPOINTED')
        else:
            status = 'NOT_STARTED'
        carrier_data[code][carrier_name] = {'status': status, 'producerNumber': None}

    # "North American" col 15 → both Life and Annuity
    val_na = clean_str(row[15]) if len(row) > 15 else None
    if val_na:
        status_na = STATUS_MAP.get(val_na.lower(), 'APPOINTED')
    else:
        status_na = 'NOT_STARTED'
    carrier_data[code]['North American Life'] = {'status': status_na, 'producerNumber': None}
    carrier_data[code]['North American Annuity'] = {'status': status_na, 'producerNumber': None}

# ── 3. Parse main "Empower Tracker" ───────────────────────────────────────
agents = []
seen_codes = set()
seen_emails = set()

ws_main = wb['Empower Tracker']
rows_main = list(ws_main.iter_rows(values_only=True))

for row in rows_main[2:]:  # skip header row
    agent_code = clean_str(row[3])
    agent_name = clean_str(row[4])
    email = clean_str(row[16])

    # Must have code and name
    if not agent_code or not agent_name:
        continue
    # Skip section header rows like "Active Before June 2024"
    if not re.match(r'^[A-Z]\d+', agent_code or ''):
        continue
    # Deduplicate by code
    if agent_code in seen_codes:
        continue
    seen_codes.add(agent_code)
    # Deduplicate by email
    if email:
        email_lower = email.lower()
        if email_lower in seen_emails:
            email = None  # will be seeded without email
        else:
            seen_emails.add(email_lower)

    raw_status = clean_str(row[10]) or 'Active'
    status = 'ACTIVE' if raw_status.lower() == 'active' else 'INACTIVE'

    first_name, last_name = split_name(agent_name)

    # Carrier appointments: start from carrier_data, then overlay producer #s from main sheet
    carriers = {}
    # Copy from carrier sheet
    for carrier_name in [
        'ANICO Life', 'ANICO Annuity', 'Augustar', 'Corebridge Life', 'Corebridge Annuity',
        'Lincoln', 'Foresters', 'Mutual of Omaha', 'SILAC', 'American Equity',
        'North American Life', 'North American Annuity', 'F&G Life', 'F&G Annuity',
        'Equitrust', 'Prudential'
    ]:
        src = (carrier_data.get(agent_code) or {}).get(carrier_name, {})
        carriers[carrier_name] = {
            'status': src.get('status', 'NOT_STARTED'),
            'producerNumber': src.get('producerNumber'),
        }

    # Overlay producer numbers from main sheet (these mean APPOINTED)
    for col_idx, (carrier_name, _) in PRODUCER_COL_MAP.items():
        if len(row) > col_idx and row[col_idx]:
            raw = clean_str(row[col_idx])
            if raw and raw.lower() not in ('false', 'true', '0', 'no'):
                # Looks like a real producer number
                prod_num = raw.split('.')[0] if raw.replace('.','').isdigit() else raw
                carriers[carrier_name] = {
                    'status': 'APPOINTED',
                    'producerNumber': prod_num,
                }

    pd = phase_data.get(agent_code, {})

    def gcol(r, i):
        return r[i] if len(r) > i else None

    agents.append({
        'agentCode': agent_code,
        'firstName': first_name,
        'lastName': last_name,
        'email': email,
        'state': clean_str(gcol(row, 1)),
        'icaDate': fmt_date(gcol(row, 2)),
        'recruiter': clean_str(gcol(row, 5)),
        'cft': clean_str(gcol(row, 6)),
        'eliteCft': clean_str(gcol(row, 7)),
        'status': status,
        'phase': pd.get('phase', 1),
        'goal': pd.get('goal'),
        'trainer': pd.get('trainer'),
        'initialPointOfContact': clean_str(gcol(row, 11)),
        'npn': clean_npn(gcol(row, 13)),
        'dateOfBirth': fmt_date(gcol(row, 14)),
        'phone': clean_str(gcol(row, 15)),
        'examDate': fmt_date(gcol(row, 27)),
        'licenseNumber': clean_str(gcol(row, 28)),
        'dateSubmittedToGfi': fmt_date(gcol(row, 29)),
        'clientProduct': clean_str(gcol(row, 24)),
        'licenseProcess': clean_str(gcol(row, 25)),
        'welcomeLetterSentAt': fmt_date(gcol(row, 17)),
        'discordJoinDate': fmt_date(gcol(row, 18)),
        'carriers': carriers,
    })

output_path = '/Users/melineeminhas/dev/allfinancialfreedom-web/scripts/agents-seed-data.json'
with open(output_path, 'w') as f:
    json.dump(agents, f, indent=2)

print(f'✓ Parsed {len(agents)} agents → {output_path}')
active = sum(1 for a in agents if a['status'] == 'ACTIVE')
with_email = sum(1 for a in agents if a['email'])
print(f'  Active: {active}  Inactive: {len(agents)-active}')
print(f'  With email: {with_email}  Without: {len(agents)-with_email}')
phase_counts = {}
for a in agents:
    p = a['phase']
    phase_counts[p] = phase_counts.get(p, 0) + 1
print(f'  Phase distribution: {dict(sorted(phase_counts.items()))}')
